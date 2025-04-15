import requests
import json
import pandas as pd
from datetime import datetime
import schedule
import time
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# Cấu hình Bitbucket API
SERVER = 'bitbucket.abc.com'
BITBUCKET_BASE_URL = f"https://{SERVER}"
PR_API_ENDPOINT = "/rest/api/1.0/projects/{project}/repos/{repo}/pull-requests/{pullRequestId}"
PR_COMMITS_ENDPOINT = "/rest/api/1.0/projects/{project}/repos/{repo}/pull-requests/{pullRequestId}/commits"
PR_ACTIVITIES_ENDPOINT = "/rest/api/1.0/projects/{project}/repos/{repo}/pull-requests/{pullRequestId}/activities"
BITBUCKET_TOKEN = "Yoru TOken"
TEAM_MEMBERS = [
    {"displayName": "Displayname 1", "username": "Displayname1@abc.com"},
    {"displayName": "Displayname 2", "username": "Displayname2@abc.com"},
]
PROJECT_REPOS = [
    {"project": "PROJECT1", "repo": "reponame1"},
    {"project": "PROJECT2", "repo": "reponame2"},  
]

# Cấu hình file Excel
EXCEL_FILE = "C:/.../.../PR_test_Auto_colect.xlsx"
SHEET_NAME = "2025"

def get_bitbucket_pull_request_details(project, repo, pullRequestId):
    """Lấy thông tin chi tiết của một Pull Request."""
    url = f"{BITBUCKET_BASE_URL}{PR_API_ENDPOINT.format(project=project, repo=repo, pullRequestId=pullRequestId)}"
    headers = {'Authorization': f'Bearer {BITBUCKET_TOKEN}'}
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"Lỗi khi gọi API chi tiết PR {pullRequestId} cho {project}/{repo}: {e}")
        return None
    except json.JSONDecodeError as e:
        print(f"Lỗi khi giải mã JSON chi tiết PR {pullRequestId} cho {project}/{repo}: {e}")
        return None

def get_bitbucket_pull_request_commits(project, repo, pullRequestId):
    """Lấy danh sách commits của một Pull Request và tính tổng additions/deletions."""
    url = f"{BITBUCKET_BASE_URL}{PR_COMMITS_ENDPOINT.format(project=project, repo=repo, pullRequestId=pullRequestId)}"
    headers = {'Authorization': f'Bearer {BITBUCKET_TOKEN}'}
    additions = 0
    deletions = 0
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        data = response.json()
        for commit in data.get('values', []):
            properties = commit.get('properties', {})
            additions += properties.get('linesAdded', 0)
            deletions += properties.get('linesRemoved', 0)
        return additions, deletions
    except requests.exceptions.RequestException as e:
        print(f"Lỗi khi gọi API commits PR {pullRequestId} cho {project}/{repo}: {e}")
        return 0, 0
    except json.JSONDecodeError as e:
        print(f"Lỗi khi giải mã JSON commits PR {pullRequestId} cho {project}/{repo}: {e}")
        return 0, 0

def get_bitbucket_pull_request_activities(project, repo, pullRequestId):
    """Lấy danh sách hoạt động của một Pull Request (approvals, comments)."""
    url = f"{BITBUCKET_BASE_URL}{PR_ACTIVITIES_ENDPOINT.format(project=project, repo=repo, pullRequestId=pullRequestId)}"
    headers = {'Authorization': f'Bearer {BITBUCKET_TOKEN}'}
    approvals = 0
    comments = 0
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        data = response.json()
        for activity in data.get('values', []):
            if activity['action'] == 'APPROVED':
                approvals += 1
            elif activity['action'] == 'COMMENTED':
                comments += 1
        return approvals, comments
    except requests.exceptions.RequestException as e:
        print(f"Lỗi khi gọi API activities PR {pullRequestId} cho {project}/{repo}: {e}")
        return 0, 0
    except json.JSONDecodeError as e:
        print(f"Lỗi khi giải mã JSON activities PR {pullRequestId} cho {project}/{repo}: {e}")
        return 0, 0


def get_bitbucket_pull_requests_list(project, repo, team_members, year=2025, expected_max_pr=None, max_requests=200):
    """Lấy danh sách Pull Requests từ Bitbucket cho team members (phân biệt theo username) trong năm cụ thể."""
    all_pull_requests_data = []
    base_url = f"{BITBUCKET_BASE_URL}/rest/api/1.0/projects/{project}/repos/{repo}/pull-requests"
    headers = {'Authorization': f'Bearer {BITBUCKET_TOKEN}'}
    params = {'limit': 500}  # Sử dụng limit 500 
    statuses = ['OPEN', 'MERGED', 'DECLINED']
    seen_next_page_starts = set()
    total_collected_pr = 0
    request_count = 0

    for status in statuses:
        url = f"{base_url}?state={status}"
        while url and request_count < max_requests:
            try:
                request_count += 1
                response = requests.get(url, headers=headers, params=params)
                response.raise_for_status()
                data = response.json()
                num_pr_in_page = len(data.get('values', []))
                total_collected_pr += num_pr_in_page
                print(f"Đã thu thập {total_collected_pr} PR cho trạng thái {status}, request #{request_count}, nextPageStart hiện tại: {data.get('nextPageStart')}")

                for pr in data.get('values', []):
                    author_info = pr['author']['user']
                    author_display_name = author_info.get('displayName')
                    author_username = author_info.get('name')  # Lấy username từ API

                    created_timestamp = pr['createdDate'] / 1000
                    created_time = datetime.fromtimestamp(created_timestamp)
                    created_date = created_time.date()

                    for member in team_members:
                        #print(f"Author Username from API: '{author_username}'")
                        if member.get('username') == author_username and created_date.year == year:
                            pr_details = get_bitbucket_pull_request_details(project, repo, pr['id'])
                            if pr_details:
                                additions, deletions = get_bitbucket_pull_request_commits(project, repo, pr['id'])
                                approvals, comments = get_bitbucket_pull_request_activities(project, repo, pr['id'])

                                closed_time = None
                                if pr_details.get('closedDate'):
                                    closed_time = datetime.fromtimestamp(pr_details['closedDate'] / 1000)

                                cycle_time_days = None
                                if created_time:
                                    if closed_time:
                                        time_difference = closed_time - created_time
                                        cycle_time_days = time_difference.total_seconds() / (60 * 60 * 24)
                                    elif pr['state'] == 'OPEN':
                                        now = datetime.now()
                                        time_difference = now - created_time
                                        cycle_time_days = time_difference.total_seconds() / (60 * 60 * 24)

                                linked_issues = [issue['key'] for issue in pr_details.get('properties', {}).get('jiraIssues', [])]

                                all_pull_requests_data.append({
                                    'Creator': author_display_name,
                                    'Username':author_username,
                                    'ID': pr['id'],
                                    'Title': pr['title'],
                                    'Created At': created_time,
                                    'Closed At': closed_time,
                                    'State': pr['state'],
                                    'From (Head Branch)': pr['fromRef']['displayId'],
                                    'To (Base Branch)': pr['toRef']['displayId'],
                                    'Base Repo': repo,
                                    'Cycle Time': cycle_time_days,
                                    'Additions': additions,
                                    'Deletions': deletions,
                                    'Approvals': approvals,
                                    'Comments': comments,
                                    'Linket Issues': ", ".join(linked_issues),
                                    'Created Date Sort': created_date
                                })
                            break  # Đã tìm thấy member phù hợp, không cần duyệt tiếp

                is_last_page = data.get('isLastPage', True)
                if is_last_page:
                    break

                next_page_url = data.get('nextPageStart')
                print(f"nextPageStart: {next_page_url}")

                if next_page_url in seen_next_page_starts and total_collected_pr > 0:
                    print(f"Warning: nextPageStart '{next_page_url}' đã được thấy trước đó. Dừng phân trang.")
                    break
                seen_next_page_starts.add(next_page_url)

                if expected_max_pr is not None and total_collected_pr >= expected_max_pr:
                    print(f"Info: Đã thu thập khoảng {expected_max_pr} PR. Dừng phân trang.")
                    break

                if next_page_url:
                    if isinstance(next_page_url, str) and (next_page_url.startswith('http') or next_page_url.startswith('/')):
                        url = next_page_url
                    elif isinstance(next_page_url, (int, str)) and str(next_page_url).isdigit():
                        url = f"{base_url}?state={status}&start={next_page_url}&limit={params['limit']}"
                    else:
                        print(f"Warning: Unhandled nextPageStart value: {next_page_url}. Stopping pagination.")
                        url = None
                else:
                    url = None

            except Exception as e:
                print(f"Lỗi trong quá trình thu thập PR ({status}): {e}")
                return []
            time.sleep(0.1)

        seen_next_page_starts = set()
        total_collected_pr_for_status = 0
        request_count_for_status = 0

    return all_pull_requests_data

def analyze_pull_requests(pull_requests_data):
    """Phân tích danh sách Pull Requests và sắp xếp theo Created At."""
    df = pd.DataFrame(pull_requests_data)
    df = df.sort_values(by='Created Date Sort', ascending=True)
    df = df.drop(columns=['Created Date Sort'])  # Xóa cột phụ sau khi sort
    return df

def update_excel(df):
    """Cập nhật dữ liệu vào file Excel."""
    if os.path.exists(EXCEL_FILE):
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        if SHEET_NAME in workbook.sheetnames:
            sheet = workbook[SHEET_NAME]
            sheet.delete_rows(1, sheet.max_row)
        else:
            sheet = workbook.create_sheet(SHEET_NAME)
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = SHEET_NAME

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    workbook.save(EXCEL_FILE)
    print(f"Dữ liệu Pull Requests năm 2025 đã được cập nhật vào file Excel: {EXCEL_FILE}, sheet: {SHEET_NAME}")

def job():
    """Công việc chính: lấy dữ liệu Pull Requests từ Bitbucket cho năm 2025 và cập nhật Excel."""
    all_pull_requests_data = []
    year_to_collect = 2025

    for repo_info in PROJECT_REPOS:
        project = repo_info['project']
        repo = repo_info['repo']
        print(f"Đang thu thập PR cho project: {project}, repository: {repo} trong năm {year_to_collect}")
        prs_data = get_bitbucket_pull_requests_list(project, repo, TEAM_MEMBERS, year_to_collect)
        all_pull_requests_data.extend(prs_data)

    if all_pull_requests_data:
        pr_df = analyze_pull_requests(all_pull_requests_data)
        update_excel(pr_df)
        print("Hoàn thành thu thập và cập nhật dữ liệu PR cho năm 2025 (đã sắp xếp theo Created At).")
    else:
        print(f"Không có Pull Requests nào được tìm thấy cho team members trong năm {year_to_collect} trên các repos đã chỉ định.")

#if __name__ == "__main__":
    #print("Script đang chạy. Dữ liệu PR năm 2025 sẽ được thu thập hàng ngày vào lúc 16:59.")
    #job() # Chạy ngay khi script được thực thi (cho test)

schedule.every().day.at("11:57").do(job)
schedule.every().day.at("16:47").do(job)

while True:
    schedule.run_pending()
    time.sleep(60)