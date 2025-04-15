import os
from jira import JIRA
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from datetime import datetime
import schedule
import time

# Cấu hình Jira
JIRA_URL = "https://jira.abc.com/"  # Thay thế bằng URL Jira của bạn
JIRA_TOKEN = "your token"  # Thay thế bằng token Jira của bạn
JIRA_QUERY = 'Your query command'  # Thay thế bằng truy vấn JQL của bạn

# Cấu hình file Excel
EXCEL_FILE = "C:/.../../..your direcotry/ABCD_Bug_found_Auto_collect.xlsx"  # Thay thế bằng đường dẫn đến file Excel của bạn
SHEET_NAME = "2025"  # Thay thế bằng tên sheet bạn muốn cập nhật

def get_all_jira_data_paginated():
    """Truy xuất tất cả dữ liệu từ Jira bằng pagination."""
    jira_options = {'server': JIRA_URL}
    jira = JIRA(options=jira_options, token_auth=JIRA_TOKEN)
    issues = []
    start_at = 0
    max_results = 1000  # Bạn có thể điều chỉnh số lượng này

    while True:
        current_issues = jira.search_issues(JIRA_QUERY, startAt=start_at, maxResults=max_results, expand='changelog')
        issues.extend(current_issues)
        if len(current_issues) < max_results:
            break  # Đã lấy hết kết quả
        start_at += max_results

    data = []
    for issue in issues:
        resolution_date = None
        if issue.fields.resolutiondate:
            resolution_date = issue.fields.resolutiondate

        story_point_value = None
        if hasattr(issue.fields, 'customfield_10002'):
            story_point_value = issue.fields.customfield_10002

        automated_tc_value = None
        if hasattr(issue.fields, 'customfield_11219'):
            if hasattr(issue.fields.customfield_11219, 'value'):
                automated_tc_value = issue.fields.customfield_11219.value
            else:
                automated_tc_value = issue.fields.customfield_11219
        
        manual_executed_raw = None
        if hasattr(issue.fields, 'customfield_11202'):
            if hasattr(issue.fields.customfield_11202, 'value'):
                manual_executed_raw = issue.fields.customfield_11202.value
            else:
                manual_executed_raw = issue.fields.customfield_11202

        manual_executed_value = None
        if manual_executed_raw is not None:
            try:
                manual_executed_value = int(manual_executed_raw)  # Cố gắng chuyển đổi sang số nguyên
            except ValueError:
                try:
                    manual_executed_value = float(manual_executed_raw) # Cố gắng chuyển đổi sang số thực
                except ValueError:
                    manual_executed_value = manual_executed_raw # Nếu không chuyển đổi được, giữ nguyên giá trị ban đầu
        
        reporter_username = issue.fields.reporter.name if issue.fields.reporter else None
        assignee_username = issue.fields.assignee.name if issue.fields.assignee else None
        
        data.append({
            "Issue Type": issue.fields.issuetype.name if issue.fields.issuetype else None,
            "Project key": issue.fields.project.key if issue.fields.project else None,
            "Issue key": issue.key,
            "Summary": issue.fields.summary,
            "Created": issue.fields.created,
            "Resolved": resolution_date,
            "Status": issue.fields.status.name if issue.fields.status else None,
            "Resolution": issue.fields.resolution.name if issue.fields.resolution else None,
            "Reporter": issue.fields.reporter.displayName if issue.fields.reporter else None,
            "Reporter_user": reporter_username,
            "Assignee": issue.fields.assignee.displayName if issue.fields.assignee else "Unassigned",
            "Assignee_user": assignee_username if issue.fields.assignee else None,
            "Story Points": story_point_value,
            "Automated TC": automated_tc_value,
            "Manual Executed": manual_executed_value,
        })
    return pd.DataFrame(data)

def update_excel(df):
    """Cập nhật dữ liệu vào file Excel."""
    if os.path.exists(EXCEL_FILE):
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook[SHEET_NAME]
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = SHEET_NAME

    # Xóa dữ liệu cũ (nếu có)
    sheet.delete_rows(1, sheet.max_row)

    # Ghi dữ liệu mới
    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    workbook.save(EXCEL_FILE)

def job():
    """Công việc chính: lấy TẤT CẢ dữ liệu Jira và cập nhật Excel."""
    print(f"Running job at {datetime.now()} (Ho Chi Minh City Time)")
    jira_data = get_all_jira_data_paginated()
    if jira_data is not None and not jira_data.empty:
        update_excel(jira_data)
    else:
        print("Không có dữ liệu Jira nào được tìm thấy.")
    print("Job resolved query finished.")

# Gọi hàm job() trực tiếp để chạy test ngay lập tức, hoặc chạy trong Task Schedule của windows
#if __name__ == "__main__":
#    job()

# Lên lịch công việc chạy mỗi ngày một lần vào lúc 8 giờ sáng
schedule.every().day.at("11:53").do(job)
schedule.every().day.at("16:43").do(job)

while True:
    schedule.run_pending()
    time.sleep(60)  # Kiểm tra mỗi phút